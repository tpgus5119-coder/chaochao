#!/usr/bin/env python3
"""GYBM 선배들이 실제로 본 **단어시험**을 전수 해부한다 → docs/senior-words.md

  python3 tools/senior_words.py              → 정리표를 화면과 문서로
  python3 tools/senior_words.py --json       → data/_senior_words.json 도 함께

왜: 22기도 같은 시험을 매주 볼 가능성이 높다. 그러면 이 낱말들이 곧 시험 범위다.
**아직 앱에 넣지 않는다**(사용자 지시). 먼저 하나도 빠짐없이 정리만 한다.

파일이 세 갈래다:
  · `20기 N회차 단어시험.xlsx`  — 회차별(거의 매일). 시트 두 장: 문제 / 답안
  · `N주차 단어 테스트(...).docx` — 주차별. kiểm tra(문제) / đáp án(답)
  · `N주차 단어시험.xlsx`         — 주차별인데 엑셀
  · `4권 베트남어 교재_단어.xlsx`  — 교재 낱말 모음(시험지가 아님)

읽는 규칙:
  · 시트 이름이 무엇이든(Kiểm tra/Đáp án./문제/답안) **베트남어 칸이 채워진 시트**를
    답안으로 본다. 답안이 없으면 문제 시트에서 한국어만 건진다.
  · 머리글 행(No. | 한글 (영어) | 베트남어)을 찾아 열 자리를 **파일마다 새로 잡는다** —
    어떤 파일은 '유형' 열이 끼어 있어 열 번호가 한 칸씩 밀린다.
  · max_row 가 94,238 같은 헛값인 파일이 있다(빈 서식이 아래로 늘어진 것).
    그래서 **빈 줄이 30줄 이어지면 끊는다.**

docx 는 python-docx 없이 읽는다 — .docx 는 zip 이고 word/document.xml 안에
표가 들어 있다. 표 칸만 뽑으면 된다.
"""
import argparse
import collections
import json
import os
import pathlib
import re
import xml.etree.ElementTree as ET
import zipfile

import openpyxl

BASE = pathlib.Path(os.path.expanduser("~/Downloads/베트남어 학습자료/선배 자료"))
GI = {"20": BASE / "20기 베트남어 단어 자료", "19": BASE / "19기 베트남어 단어 자료"}
R = pathlib.Path(__file__).resolve().parent.parent
W = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"

VI = re.compile(r"[ăâđêôơưàáảãạèéẻẽẹìíỉĩịòóỏõọùúủũụỳýỷỹỵ]", re.I)
KO = re.compile(r"[가-힣]")


def keyof(ko):
    """낱말을 견줄 수 있게 다듬는다.

    같은 낱말이 파일마다 다르게 적혀 있다:
        회차(xlsx) `위치 (location)` · 주차(docx) `위치 location` · `알다 / to know`
    글자 그대로 견주면 전부 남남이 된다 — 처음에 주차 절반이 '겹치는 회차 없음'으로
    나왔던 까닭이다. **한국어 부분만** 남겨 견준다.
    """
    s = re.sub(r"[(（][^)）]*[)）]", " ", ko)      # 괄호 안(영어 뜻) 걷어내기
    s = re.sub(r"[/·,]", " ", s)
    s = "".join(ch for ch in s if KO.search(ch) or ch in " ~")
    return re.sub(r"\s+", " ", s).strip()


def blank(v):
    return v is None or not str(v).strip()


def rows_of(ws, cap_blank=30):
    """헛되게 늘어진 빈 줄에서 끊고 읽는다."""
    out, run = [], 0
    for r in ws.iter_rows(values_only=True):
        cells = ["" if c is None else str(c).strip() for c in r]
        if not any(cells):
            run += 1
            if run >= cap_blank:
                break
            continue
        run = 0
        out.append(cells)
    return out


def header_map(rows):
    """머리글 행을 찾아 **표 한 벌마다** {열이름: 열번호} 를 만들어 목록으로 돌려준다.

    한 줄에 표가 여러 벌 나란히 놓인 것이 있다 — 19기 토요시험이 그렇다:
        No.|단어|뜻 | No.|단어|뜻 | No.|단어|뜻 | No.|단어|뜻
    'No.' 가 새로 나오면 새 표가 시작된 것으로 본다. 한 벌만 읽으면 100낱말 중
    25개만 건진다.
    """
    for i, r in enumerate(rows[:12]):
        blocks, cur = [], {}
        for c, v in enumerate(r):
            s = v.replace(" ", "")
            if s.lower().startswith("no") or s == "번호":
                if cur:
                    blocks.append(cur)
                cur = {"no": c}
            elif "한글" in s or "한국" in s or s == "뜻":
                # 3·4회차 답안 시트는 열이 뒤집혀 있다:
                #   No. | 유형 | 베트남어 | 영어 | 뜻   ← 한국어가 '뜻' 이라는 이름으로 맨 끝
                # '한글' 만 찾다가 이 두 회차의 답을 통째로 놓쳤다.
                # **먼저 나온 것을 잡는다**(setdefault) — 19기 94회차 머리글은
                #   No. | 유형 | 한글 (영어) | 베트남어 | | 한국어 | 영어
                # 처럼 한국어 이름이 둘이라, 나중 것으로 덮으면 빈 열을 가리켜
                # 그 회차 낱말을 통째로 잃는다.
                cur.setdefault("ko", c)
            elif "베트남" in s or s == "단어":
                # 19기는 열 이름이 `No.|뜻|단어` 다 — '단어' 가 곧 베트남어 칸이다.
                # 이걸 안 넣으면 19기 파일 절반에서 답이 통째로 빠진다.
                cur.setdefault("vi", c)
            elif "유형" in s:
                cur["kind"] = c
        if cur:
            blocks.append(cur)
        blocks = [b for b in blocks if "ko" in b]
        if blocks:
            return i, blocks
    return -1, []


def not_viet(got, need=20, floor=.25):
    """답이 채워져 있는데 **베트남어가 아닌** 시트인가.

    베트남어는 성조·모음 부호가 촘촘해 낱말 스무 개 중 넷도 안 걸리는 일이 없다.
    """
    ans = [v for _, v, _ in got if v]
    return len(ans) >= need and sum(1 for v in ans if VI.search(v)) / len(ans) < floor


def harvest(rows, hi, blocks):
    """머리글 아래를 훑어 (한국어, 베트남어, 유형) 을 모은다.

    표가 여러 벌이면 **한 벌씩 차례로** 읽는다 — 줄을 가로질러 읽으면
    1·26·51·76 순서가 되어 원래 번호 차례가 흐트러진다.
    """
    got = []
    for j in blocks:
        for r in rows[hi + 1:]:
            cell = lambda k: (r[j[k]] if k in j and j[k] < len(r) else "")   # noqa: B023
            ko, vi, kind = cell("ko"), cell("vi"), cell("kind")
            if blank(ko):
                continue
            if not KO.search(ko) and not re.search(r"[A-Za-z]", ko):
                continue                       # 안내문·빈 서식 줄
            got.append((ko.strip(), (vi or "").strip(), (kind or "").strip()))
    return got


TITLE_NO = re.compile(r"(\d+)\s*회차\s*단어")


def sheets_of(f):
    """엑셀 한 권을 **시트마다 따로** 읽는다 — 한 권에 여러 회차가 들어 있기 때문.

    처음에 한 권을 한 덩어리로 뭉쳐 읽어 8·9·10회차를 '없다'고 잘못 보고했다.
    실제로는 `8주차 단어시험.xlsx` 시트가 ['4회차'..'8회차'] 다섯 장이었고,
    `9주차`·`10주차` 파일은 이름만 주차일 뿐 시트 첫 줄이 '9회차'·'10회차'였다.
    시트 이름과 시트 첫 줄에서 회차를 읽고, 없으면 파일 이름을 따른다.
    """
    try:
        wb = openpyxl.load_workbook(f, data_only=True, read_only=True)
    except Exception as e:
        return [], f"못 열었다: {e}"
    out = []
    for ws in wb.worksheets:
        rows = rows_of(ws)
        if not rows:
            continue
        no = None
        m = TITLE_NO.search(ws.title) or re.search(r"(\d+)\s*회차", ws.title)
        if m:
            no = int(m.group(1))
        if no is None:
            for r in rows[:3]:
                for v in r:
                    mm = TITLE_NO.search(v or "")
                    if mm:
                        no = int(mm.group(1))
                        break
                if no:
                    break
        hi, blocks = header_map(rows)
        if hi < 0:
            continue
        got = harvest(rows, hi, blocks)
        if not_viet(got):
            # 19기 토요시험 서식은 **인도네시아 GYBM 것을 가져다 쓴 것**이라
            # `Sheet1_입력` 시트에 인도네시아어 보기가 그대로 남아 있다
            # (`소비재 → barang konsumsi`, `표현하다 → Ungkapkan`).
            # 이걸 안 걸러내면 주간 시험 낱말이 곱절로 부풀고 절반이 딴 나라 말이 된다.
            continue
        if got:
            out.append((no, got, ws.title))
    wb.close()
    return out, ""


def from_xlsx(f):
    """엑셀 한 권에서 (한국어, 베트남어, 유형) 목록을 뽑는다. 모든 시트를 본다.

    시트 첫 줄에 'N회차 단어 테스트' 가 적혀 있으면 **그 번호를 파일 이름보다 믿는다** —
    `20기 62회차 단어시험.xlsx` 의 문제 시트가 '61회차'라고 적고 있었다(파일 이름이 틀렸다).
    """
    try:
        wb = openpyxl.load_workbook(f, data_only=True, read_only=True)
    except Exception as e:
        return [], f"못 열었다: {e}", None
    best, note, said_no = [], "", None
    for ws in wb.worksheets:
        rows = rows_of(ws)
        for r in rows[:3]:
            for v in r:
                m = TITLE_NO.search(v or "")
                if m and said_no is None:
                    said_no = int(m.group(1))
        hi, blocks = header_map(rows)
        if hi < 0:
            continue
        got = harvest(rows, hi, blocks)
        # 베트남어가 채워진 시트가 답안지다 — 그쪽을 택한다
        score = sum(1 for _, v, _ in got if v)
        if score > sum(1 for _, v, _ in best if v) or (not best and got):
            best, note = got, ws.title
    wb.close()
    return best, note, said_no


def from_docx(f):
    """python-docx 없이 표 칸만 뽑는다 — .docx 는 zip 이고 안에 xml 이 있다."""
    try:
        with zipfile.ZipFile(f) as z:
            xml = z.read("word/document.xml")
    except Exception as e:
        return [], f"못 열었다: {e}", None
    root = ET.fromstring(xml)
    out = []
    for tr in root.iter(f"{W}tr"):
        cells = []
        for tc in tr.iter(f"{W}tc"):
            txt = "".join(t.text or "" for t in tc.iter(f"{W}t")).strip()
            cells.append(txt)
        if len(cells) >= 2:
            out.append(cells)
    hi, blocks = header_map(out)
    if hi >= 0:
        return harvest(out, hi, blocks), "머리글 있는 표", None

    # 주차 테스트 docx 는 **머리글이 아예 없다.** 한 줄에 낱말이 넷씩 나란히 들어 있다:
    # [한국어, 베트남어, 한국어, 베트남어, 한국어, 베트남어, 한국어, 베트남어]
    # 처음엔 머리글을 못 찾아 36개 파일을 통째로 버렸다 — 짝으로 훑어야 읽힌다.
    got = []
    for r in out:
        for a, b in zip(r[0::2], r[1::2]):
            a, b = (a or "").strip(), (b or "").strip()
            if not a or not KO.search(a):
                continue                       # 한국어가 없으면 낱말 칸이 아니다
            if b and KO.search(b) and not VI.search(b):
                continue                       # 오른쪽도 한국어면 짝이 어긋난 것
            got.append((a, b, ""))
    return ((got, "머리글 없는 넷씩 표", None) if got
            else ([], "표는 있는데 낱말을 못 찾았다", None))


def from_pdf(f):
    """PDF 시험지에서 낱말을 뽑는다.

    처음엔 PDF 를 아예 안 읽었다 — 그 바람에 **2회차를 통째로 놓쳤다**(PDF 로만 있다).
    꼴: `1 일본어 japanese language` 처럼 번호 + 한국어 + 영어. 답(베트남어)은
    문제지 PDF 에 없으므로 한국어만 건진다.
    """
    try:
        from pypdf import PdfReader
        r = PdfReader(f)
        txt = "\n".join((pg.extract_text() or "") for pg in r.pages)
    except Exception as e:
        return [], f"못 열었다: {e}", None
    got = []
    for line in txt.split("\n"):
        m = re.match(r"^\s*(\d{1,3})[.)]?\s+(.+)$", line.strip())
        if not m:
            continue
        no, rest = int(m.group(1)), m.group(2).strip()
        if not (1 <= no <= 200) or not KO.search(rest):
            continue
        got.append((rest, "", ""))
    return (got, "PDF", None) if got else ([], "PDF 인데 낱말 줄을 못 찾았다", None)


def label(name):
    """파일 이름에서 (갈래, 번호) 를 읽는다.

    **두 시험이 섞여 있다.** 처음에 이걸 못 갈라 docx `16회차 단어 테스트` 와
    xlsx `20기 16회차 단어시험` 을 같은 것으로 합쳐 버렸다 — 다른 시험이다.
      · `20기 N회차 단어시험`  = 날마다 보는 시험 (30낱말)
      · `N주차/N회차 단어 테스트` = 주마다 보는 시험 (100낱말)
        — 이 묶음은 1~15는 '주차', 16~20은 '회차'로 이름이 바뀌었을 뿐 한 줄기다.
      · `N주차 단어시험.xlsx`   = 주마다 보는 시험(엑셀판)
    """
    n = re.sub(r"^TalkFile_", "", name).replace(".xlsx.xlsx", ".xlsx")
    # 19기는 토요일에 주간 시험을 봤다: `토요시험단어 1회차 250627` · `250712 3차 주간단어`.
    # 날짜(6/27 → 7/5 → 7/12)가 한 주 간격이라 한 줄기다.
    if "토요시험" in n or "주간단어" in n:
        m = re.search(r"(\d+)\s*회?차", n)
        return ("주간", int(m.group(1))) if m else ("기타", 0)
    if re.match(r"\s*\d+\s*기", n):            # `19기 …` `20기 …` = 날마다 보는 시험
        # `20기 ( 5 )회차` 처럼 괄호에 싸인 것 · 19기 `단어시험24회차` 처럼 붙여 쓴 것 둘 다
        m = re.search(r"\(?\s*(\d+)\s*\)?\s*회차", n)
        if m:
            return "일일", int(m.group(1))
        m = re.search(r"단어시험\s*(\d+)", n)
        if m:
            return "일일", int(m.group(1))
        return "기타", 0
    m = re.search(r"(\d+)\s*주차", n)
    if m:
        return "주간", int(m.group(1))
    m = re.search(r"(\d+)\s*회차", n)          # 주간 묶음의 16~20 (이름만 '회차')
    if m:
        return "주간", int(m.group(1))
    return "기타", 0


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--json", action="store_true")
    ap.add_argument("--gi", default="20", choices=sorted(GI), help="기수 (20 또는 19)")
    a = ap.parse_args()
    SRC = GI[a.gi]
    suffix = "" if a.gi == "20" else f"-{a.gi}"
    if not SRC.exists():
        raise SystemExit(f"자료가 없다: {SRC}")

    files = sorted(p for p in SRC.iterdir()
                   if p.suffix.lower() in (".xlsx", ".docx", ".pdf")
                   and not p.name.startswith("~$"))
    recs = {}                                   # (갈래, 번호) → {ko: vi}
    per_file, fails, conflicts = [], [], []
    for f in files:
        ext = f.suffix.lower()
        kind, no = label(f.name)
        if ext == ".xlsx":
            # 시트마다 회차가 다를 수 있다 — 시트 단위로 담는다
            sh, err = sheets_of(f)
            if not sh:
                fails.append((f.name, err or "시트에서 낱말을 못 찾았다"))
                continue
            # 시트가 저마다 다른 회차를 대면 '한 권에 여러 회차'인 파일이다
            nos = {s[0] for s in sh if s[0]}
            multi = len(nos) > 1
            for sno, got, title in sh:
                # 파일 이름이 'N주차'라도 시트가 'N회차 단어 테스트'라고 밝히면 일일이다.
                # `9주차 단어시험.xlsx` 의 시트 첫 줄이 '9회차 단어 테스트' 였다 —
                # 이름만 주차일 뿐 내용은 일일 시험이라, 이름만 믿어 주간으로 넣었던 것을
                # 바로잡는다. (주간은 100낱말, 일일은 30낱말이라 크기로도 갈린다.)
                if multi and sno:
                    k2 = ("일일", sno)
                elif kind == "주간" and sno and len(got) <= 45:
                    k2 = ("일일", sno)
                else:
                    k2 = (kind, no)
                if not multi and sno and sno != no and kind == "일일":
                    conflicts.append((f.name, no, sno))
                bag = recs.setdefault(k2, {})
                for ko, vi, _ in got:
                    if ko not in bag or (vi and not bag[ko]):
                        bag[ko] = vi
                per_file.append((f.name, k2[0], k2[1], len(got),
                                 sum(1 for _, v, _ in got if v), title))
            continue
        got, note, said = (from_docx(f) if ext == ".docx" else from_pdf(f))
        # 시트 첫 줄의 'N회차'는 **믿지 않는다.** 7개 파일에서 파일 이름과 어긋났는데,
        # 지난 회차 파일을 복사해 제목만 안 고친 흔적이다(`7회차` 파일에 `81회차`라고
        # 적혀 있고, 한 파일 안에서 문제 시트와 답안 시트가 서로 다른 번호를 대기도 한다).
        # 파일 이름을 믿되, 어긋난 것은 문서 끝에 적어 대표님이 볼 수 있게 남긴다.
        if said and kind == "일일" and said != no:
            conflicts.append((f.name, no, said))
        if not got:
            fails.append((f.name, note))
            continue
        key = (kind, no)
        bag = recs.setdefault(key, {})
        for ko, vi, _ in got:
            # 답이 있는 쪽이 이긴다 — 문제지와 답안지가 같은 회차로 들어오기 때문
            if ko not in bag or (vi and not bag[ko]):
                bag[ko] = vi
        per_file.append((f.name, kind, no, len(got), sum(1 for _, v, _ in got if v), note))

    L = []
    L.append(f"# GYBM {a.gi}기 선배 단어시험 — 전수 정리")
    L.append("")
    L.append(f"자료: `{SRC}` · 파일 {len(files)}개(xlsx {sum(1 for f in files if f.suffix=='.xlsx')} ·"
             f" docx {sum(1 for f in files if f.suffix=='.docx')} ·"
             f" pdf {sum(1 for f in files if f.suffix=='.pdf')})")
    L.append("")
    L.append("> 아직 앱에 넣지 않았다. 정리만 한 것이다.")
    L.append("")

    weeks = sorted(k for k in recs if k[0] == "주간")
    rounds = sorted(k for k in recs if k[0] == "일일")
    etc = sorted(k for k in recs if k[0] == "기타")

    L.append("## 한눈에")
    L.append("")
    L.append("| 갈래 | 몇 개 | 번호 범위 | 낱말(중복 포함) | 낱말(서로 다른 것) |")
    L.append("|---|---:|---|---:|---:|")
    for nm, ks in (("주간 시험", weeks), ("일일 시험", rounds), ("그 밖", etc)):
        if not ks:
            continue
        allw = [w for k in ks for w in recs[k]]
        nos = [k[1] for k in ks if k[1]]
        L.append(f"| {nm} | {len(ks)} | {min(nos) if nos else '-'}~{max(nos) if nos else '-'} "
                 f"| {len(allw):,} | {len(set(allw)):,} |")
    every = [w for k in recs for w in recs[k]]
    L.append(f"| **합** | {len(recs)} | | **{len(every):,}** | **{len(set(every)):,}** |")
    L.append("")

    for nm, ks in (("주간 시험 (100낱말)", weeks), ("일일 시험 (30낱말)", rounds), ("그 밖의 파일", etc)):
        if not ks:
            continue
        L.append(f"## {nm}")
        L.append("")
        L.append("| 번호 | 낱말 수 | 답 있는 것 | 첫 낱말 다섯 |")
        L.append("|---:|---:|---:|---|")
        for k in ks:
            bag = recs[k]
            has = sum(1 for v in bag.values() if v)
            head = " · ".join(list(bag)[:5])
            L.append(f"| {k[1] or '—'} | {len(bag)} | {has} | {head[:70]} |")
        L.append("")

    miss = sorted(set(range(1, max([k[1] for k in rounds], default=0) + 1))
                  - {k[1] for k in rounds})
    if miss:
        L.append(f"**빠진 일일 회차**: {', '.join(map(str, miss))}")
        L.append("")
    if conflicts:
        L.append("## 파일 이름과 시트 제목이 어긋난 것 (파일 이름을 믿었다)")
        L.append("")
        L.append("지난 회차 파일을 복사해 제목만 안 고친 흔적으로 보인다.")
        L.append("")
        L.append("| 파일 | 파일 이름 | 시트 제목 |")
        L.append("|---|---:|---:|")
        for n, fno, said in conflicts:
            L.append(f"| `{n}` | {fno} | {said} |")
        L.append("")

    if fails:
        L.append("## 못 읽은 파일")
        L.append("")
        for n, why in fails:
            L.append(f"- `{n}` — {why}")
        L.append("")

    # 주차 시험은 그 주의 회차 시험을 모은 것인가 — 낱말 겹침으로 확인한다.
    # 표기가 파일마다 달라(`위치 (location)` vs `위치 location`) 정규화해서 견준다.
    NK = {k: {keyof(w) for w in recs[k] if keyof(w)} for k in recs}
    wkk = {k[1]: NK[k] for k in weeks}
    rdk = {k[1]: NK[k] for k in rounds}
    if wkk and rdk:
        L.append("## 주간 시험은 그 주의 일일 시험을 모은 것이다")
        L.append("")
        L.append("낱말이 파일마다 다르게 적혀 있어(`위치 (location)` · `위치 location` ·")
        L.append("`알다 / to know`) **한국어 부분만 남겨** 견줬다. 그러자 관계가 드러났다.")
        L.append("")
        L.append("| 주간 | 낱말 | 그 주의 일일 회차 | 주간 낱말이 일일로 설명되는 비율 |")
        L.append("|---:|---:|---|---:|")
        used = set()
        for w in sorted(wkk):
            hits = sorted(r for r in rdk if len(wkk[w] & rdk[r]) / max(1, len(rdk[r])) >= .5)
            used |= set(hits)
            fr = set().union(*[rdk[r] for r in hits]) if hits else set()
            pct = len(wkk[w] & fr) / max(1, len(wkk[w])) * 100
            L.append(f"| {w} | {len(wkk[w])} | {', '.join(map(str, hits)) or '—'} | {pct:.0f}% |")
        L.append("")
        L.append(f"주간에 묶인 일일 회차 {len(used)}개 / 일일 시험 {len(rdk)}개. "
                 f"주간 시험 하나가 **일일 4~5회차**를 담는다 — **일일은 날마다, 주간은 그 주 몰아서**다.")
        L.append("")
        L.append(f"정규화해서 세면 서로 다른 낱말은 **{len(set().union(*NK.values())):,}개**"
                 f"(글자 그대로 세면 {len(set(every)):,}개로 부풀어 보인다).")
        L.append("")

    # 겹침 — 같은 낱말이 몇 번 나왔나
    cnt = collections.Counter(every)
    rep = [(w, c) for w, c in cnt.most_common(20) if c > 1]
    if rep:
        L.append("## 여러 번 나온 낱말 (반복 출제)")
        L.append("")
        L.append(" · ".join(f"{w}({c})" for w, c in rep))
        L.append("")

    out = "\n".join(L)
    (R / "docs" / f"senior-words{suffix}.md").write_text(out + "\n", encoding="utf-8")
    print(out)
    print(f"\n→ docs/senior-words{suffix}.md")

    if a.json:
        j = {"note": f"GYBM {a.gi}기 선배 단어시험 전수. 앱에 아직 안 넣음.",
             "sets": [{"kind": k[0], "no": k[1],
                       "words": [{"ko": ko, "vi": vi} for ko, vi in recs[k].items()]}
                      for k in sorted(recs)]}
        p = R / "data" / f"_senior_words{suffix}.json"
        p.write_text(json.dumps(j, ensure_ascii=False, indent=1), encoding="utf-8")
        print(f"→ {p}")


if __name__ == "__main__":
    main()
