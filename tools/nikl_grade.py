#!/usr/bin/env python3
"""국립국어원 **등급별 어휘 12,019**를 쓰기 좋은 표로 바꾼다.

무엇이 달라지나 — 지금까지 쓰던 「한국어 학습용 어휘 목록」(2003, 5,965)보다
  · 두 배 넘게 크고 (초급 2,247 · 중급 4,569 · 고급 5,203)
  · **길잡이말**이 있다. 동형이의어를 갈라 주는 칸이다.
    '경우'가 '도리'인지 '어떤 형편'인지, '주문'이 注文인지 呪文인지 여기서 갈린다.
    우리 베트남어 뜻풀이가 바로 그 자리에서 틀려 있었다(경우→đạo lý, 주문→câu thần chú).
  · **유의어·반의어**가 있다. 오답지를 "같은 등급·같은 품사"로만 고르던 것을
    "뜻이 가깝되 답은 아닌 것"으로 올릴 수 있다.
  · **주제·기능**과 **의미범주**가 있다. 모의고사 주제 분포를 맞추는 잣대다.

출처: 「한국어 교육 어휘 내용 개발(4단계)」(국립국어원, 2015, 연구책임자 강현화)
      부록 + '어휘 정보.xlsx'. 원본은 ~/Documents/시험기출자료고/국립국어원-어휘.
      저장소에는 **우리가 쓰는 칸만** 추린 표를 둔다(원본 xlsx는 넣지 않는다).

실행: python3 tools/nikl_grade.py      → tools/nikl_12019.tsv
"""
import pathlib
import re
import sys

SRC = (pathlib.Path.home() / "Documents" / "시험기출자료고" / "국립국어원-어휘"
       / "어휘내용개발4단계_어휘정보.xlsx")
OUT = pathlib.Path(__file__).resolve().parent / "nikl_12019.tsv"

# 우리가 쓰는 칸만. 나머지(연어·관용어·상위어·큰말/작은말 등)는 지금 쓸 데가 없다.
COLS = ["등급", "어휘", "품사", "길잡이말", "유의어1\n(기초사전)", "반의어1\n(기초사전)",
        "주제·기능", "대범주", "소범주"]
HEAD = ["등급", "어휘", "품사", "길잡이말", "유의어", "반의어", "주제", "대범주", "소범주"]


def clean(v):
    """줄바꿈은 세로줄로, 앞뒤 공백은 지운다. 빈 칸은 빈 글자."""
    if v is None:
        return ""
    s = str(v).strip()
    if s in ("없음", "-", "None"):
        return ""
    return re.sub(r"\s*\n\s*", "|", s)


def base(word):
    """'-가02' · '가격03' 처럼 붙은 동형이의어 번호를 떼고 표제어만 남긴다.

    번호는 사전 안에서 같은 글자를 가르는 표시다. 우리는 **번호는 버리고
    길잡이말로 가른다** — 학습자에게 '가격03'을 보여 줄 수는 없으니까.
    """
    return re.sub(r"\d+$", "", word.strip().lstrip("-")).strip()


def main():
    try:
        import openpyxl
    except ImportError:
        sys.exit("openpyxl 이 필요하다:  python3 -m pip install --user openpyxl")
    if not SRC.exists():
        sys.exit(f"원본이 없다: {SRC}\n(tools/gather_archive.py 나 손으로 먼저 받는다)")

    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    # 머리글이 **두 줄**이다. 윗줄은 묶음 이름('관련어'), 아랫줄이 진짜 칸 이름
    # ('유의어1(기초사전)'). 아랫줄이 있으면 그것을, 없으면 윗줄을 쓴다.
    top = [clean(h) for h in next(rows)]
    sub = [clean(h) for h in next(rows)]
    hdr = [s or t for s, t in zip(sub, top)]
    idx = {}
    for want, name in zip(COLS, HEAD):
        w = clean(want)
        if w in hdr:
            idx[name] = hdr.index(w)
    missing = [n for n in HEAD if n not in idx]
    if missing:
        sys.exit(f"표에 없는 칸: {missing}\n실제 머리글: {hdr}")

    out, seen = [], set()
    for r in rows:
        g = clean(r[idx["등급"]])
        if g not in ("초급", "중급", "고급"):
            continue
        rec = {n: clean(r[idx[n]]) for n in HEAD}
        rec["어휘"] = base(rec["어휘"])
        if not rec["어휘"]:
            continue
        # 같은 낱말이 길잡이말만 다르게 여러 번 나온다(동형이의어) — 전부 남긴다.
        key = (rec["어휘"], rec["품사"], rec["길잡이말"])
        if key in seen:
            continue
        seen.add(key)
        out.append(rec)

    with OUT.open("w", encoding="utf-8") as f:
        f.write("\t".join(HEAD) + "\n")
        for rec in out:
            f.write("\t".join(rec[n] for n in HEAD) + "\n")

    from collections import Counter
    c = Counter(r["등급"] for r in out)
    homo = Counter(r["어휘"] for r in out)
    n_homo = sum(1 for w, k in homo.items() if k > 1)
    print(f"{OUT.name} — {len(out)}줄 ({dict(c)})")
    print(f"  길잡이말 있는 줄 {sum(1 for r in out if r['길잡이말'])}")
    print(f"  유의어 있는 줄 {sum(1 for r in out if r['유의어'])} · "
          f"반의어 {sum(1 for r in out if r['반의어'])}")
    print(f"  **한 낱말이 뜻 여럿을 가진 것 {n_homo}개** — 여기가 오역이 나던 자리다")


if __name__ == "__main__":
    main()
